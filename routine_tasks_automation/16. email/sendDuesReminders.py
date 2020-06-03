"""
sendDuesReminders.py - Рассылает сообщения на основании
    сведений из электронной таблицы об уплате взносов.
"""

import openpyxl
import smtplib
import sys

path = './routine_tasks_automation/16. email/'
wb = openpyxl.load_workbook(path + 'duesRecord.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

unpaidMemvers = []
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMemvers[name] = email

smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_address@gmail.com', sys.argv[1])
