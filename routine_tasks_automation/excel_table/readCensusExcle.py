"""
readCensusExcel.py - Формирует таблицу данных о численности
    населения и количестве переписных районов в каждом округе.
"""

from openpyxl import load_workbook
import pprint


print('Открытие рабочей книги...')
wb = load_workbook(
    filename='routine_tasks_automation/excel_table/censuspopdata.xlsx',
    data_only=True
)
sheet = wb['Лист1']
countryData = {}

print('Чтение строк...')
for row in range(2, sheet.max_row + 1):
    state = sheet['C' + str(row)].value
    country = sheet['D' + str(row)].value
    pop = sheet['E' + str(row)].value

    countryData.setdefault(state, {})
    countryData[state].setdefault(country, {'tracts': 0,
                                            'pop': 0})
    countryData[state][country]['tracts'] += 1
    countryData[state][country]['pop'] += int(pop)

print('Запись результатов...')
with open('census2010.py', 'w') as resultFile:
    resultFile.write('allData = ' + pprint.pformat(countryData))
print('Готово.')
